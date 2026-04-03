from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import OrderedDict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL_PATH = ROOT_DIR / '\u5267\u60c5.xlsx'
DEFAULT_OUTPUT_PATH = ROOT_DIR / 'js' / 'story' / 'story-generated-data.js'
STORY_ASSET_ROOT = ROOT_DIR / 'Texture' / 'story'
PORTRAIT_FOLDER = '\u7acb\u7ed8'
PORTRAIT_DIR = STORY_ASSET_ROOT / PORTRAIT_FOLDER

HEADER_ORDER = '\u5e55'
HEADER_SPEAKER = '\u4eba\u7269'
HEADER_TEXT = '\u5bf9\u8bdd'
HEADER_SCENE = '\u573a\u666f'
HEADER_EXTRA = '\u8865\u5145'
HEADER_EXTRA_TYPE = '\u8865\u5145\u7c7b\u578b'

NARRATION_SPEAKER = '\u65c1\u767d'
PROLOGUE_TITLE = '\u5e8f\u7ae0'
BACKGROUND_KEYWORD = '\u80cc\u666f'
FULL_WIDTH_COLON = '\uff1a'
FULL_WIDTH_COMMA = '\uff0c'
STRONG_CHAR = '\u5f3a'
SOFT_CHAR = '\u8f7b'
LEFT_TAG = '\u5de6'
RIGHT_TAG = '\u53f3'

SHAKE_KEYWORDS = ('\u9707\u52a8', '\u6296\u52a8')
ZOOM_KEYWORDS = (
    '\u955c\u5934\u62c9\u8fd1',
    '\u62c9\u8fd1\u955c\u5934',
    '\u62c9\u8fd1',
    '\u7279\u5199',
)
PAN_LEFT_TO_RIGHT_KEYWORDS = ('\u4ece\u5de6\u5230\u53f3', '\u5de6\u5230\u53f3')
PAN_RIGHT_TO_LEFT_KEYWORDS = ('\u4ece\u53f3\u5230\u5de6', '\u53f3\u5230\u5de6')
BACKGROUND_MAIN_VALUES = (
    '\u4e3b\u754c\u9762',
    '\u4e3b\u9875',
    '\u4e3b\u83dc\u5355',
    '\u52a8\u7269\u56ed\u4e3b\u9875',
    '\u52a8\u7269\u56ed\u4e3b\u754c\u9762',
    'zoo',
    'zoohome',
    'zoo-home',
    'mainmenu',
    'menu',
)

SUPPLEMENT_TYPE_PAN_RIGHT_TO_LEFT = 1
SUPPLEMENT_TYPE_SHAKE = 2
SUPPLEMENT_TYPE_ZOOM_IN = 3
SUPPLEMENT_TYPE_DUAL_DIALOGUE = 4
SUPPLEMENT_TYPE_ITEM_REWARD = 5
SUPPLEMENT_TYPE_ZOO_HOME = 6
SUPPLEMENT_TYPE_INTERACTION = 7
SUPPLEMENT_TYPE_OPTION = 8
SUPPLEMENT_TYPE_COLLECTION = 9


COLLECTION_SPECIES_MAP = {
    '小熊猫': {'speciesId': 'red-panda', 'speciesName': '小熊猫'},
    '水豚': {'speciesId': 'capybara', 'speciesName': '水豚'},
    '大熊猫': {'speciesId': 'giant-panda', 'speciesName': '大熊猫'},
    '雪豹': {'speciesId': 'snow-leopard', 'speciesName': '雪豹'},
    '考拉': {'speciesId': 'koala', 'speciesName': '考拉'},
    '帝企鹅': {'speciesId': 'emperor-penguin', 'speciesName': '帝企鹅'},
    '环尾狐猴': {'speciesId': 'ring-tailed-lemur', 'speciesName': '环尾狐猴'},
    '斑马': {'speciesId': 'zebra', 'speciesName': '斑马'},
    '长颈鹿': {'speciesId': 'giraffe', 'speciesName': '长颈鹿'},
    '亚洲象': {'speciesId': 'asian-elephant', 'speciesName': '亚洲象'},
    '水獭': {'speciesId': 'otter', 'speciesName': '水獭'}
}


@dataclass
class ImportResult:
    stories: dict[str, Any]
    warnings: list[str]
    story_count: int
    beat_count: int


def normalize_text(value: Any) -> str:
    if value is None:
        return ''
    return str(value).strip()


def split_extra_parts(extra_text: Any) -> list[str]:
    extra = normalize_text(extra_text)
    if not extra:
        return []
    return [part for part in re.split(r'[\s,\uFF0C\u3001/|\uFF1B;]+', extra) if part]


def parse_option_choices(extra_text: Any) -> list[dict[str, Any]] | None:
    normalized = normalize_text(extra_text)
    if not normalized:
        return None

    parts = split_extra_parts(normalized)
    if not parts:
        return None

    choices: list[dict[str, Any]] = []
    for part in parts:
        match = re.match(r'^(?P<id>\d+)-(?P<label>.+)$', part.strip())
        if not match:
            continue

        choice_id = match.group('id')
        label = normalize_text(match.group('label'))
        if not label:
            continue

        try:
            jump = int(choice_id)
        except ValueError:
            jump = 0

        choices.append({
            'id': choice_id,
            'label': label,
            'jump': jump
        })

    if not choices and len(parts) == 1:
        label = normalize_text(parts[0])
        if label:
            return [{
                'id': '1',
                'label': label,
                'jump': 1
            }]

    return choices or None


def get_option_choices(supplement_type: int, extra_text: Any) -> list[dict[str, Any]] | None:
    parsed_choices = parse_option_choices(extra_text)
    if supplement_type == SUPPLEMENT_TYPE_OPTION:
        return parsed_choices
    return parsed_choices


def stable_id(prefix: str, value: str) -> str:
    digest = hashlib.md5(value.encode('utf-8')).hexdigest()[:10]
    return f'{prefix}-{digest}'


def build_story_id(sheet_name: str) -> str:
    return 'prologue' if sheet_name == PROLOGUE_TITLE else sheet_name


def resolve_existing_asset_path(directory: Path, base_name: str) -> Path | None:
    normalized_base_name = normalize_text(base_name).replace('\\', '/').strip()
    if not normalized_base_name:
        return None

    direct_candidate = directory / normalized_base_name
    if direct_candidate.suffix:
        return direct_candidate if direct_candidate.exists() else None

    for extension in ('.webp', '.png'):
        asset_path = directory / f'{normalized_base_name}{extension}'
        if asset_path.exists():
            return asset_path

    return None


def build_story_asset_src(sheet_name: str, asset_name: str) -> str:
    normalized_asset_name = normalize_text(asset_name).replace('\\', '/').strip()
    if not normalized_asset_name:
        return ''

    asset_directory = STORY_ASSET_ROOT / sheet_name
    asset_path = resolve_existing_asset_path(asset_directory, normalized_asset_name)
    if asset_path:
        return f'./Texture/story/{sheet_name}/{asset_path.name}'

    if re.search(r'\.[A-Za-z0-9]+$', normalized_asset_name):
        return f'./Texture/story/{sheet_name}/{normalized_asset_name}'
    return f'./Texture/story/{sheet_name}/{normalized_asset_name}.png'


def build_portrait_asset_src(portrait_label: str) -> str:
    normalized_portrait_label = normalize_text(portrait_label).replace('\\', '/').strip()
    if not normalized_portrait_label:
        return ''

    portrait_path = resolve_existing_asset_path(PORTRAIT_DIR, normalized_portrait_label)
    if portrait_path:
        return f'./Texture/story/{PORTRAIT_FOLDER}/{portrait_path.name}'

    if re.search(r'\.[A-Za-z0-9]+$', normalized_portrait_label):
        return f'./Texture/story/{PORTRAIT_FOLDER}/{normalized_portrait_label}'
    return f'./Texture/story/{PORTRAIT_FOLDER}/{normalized_portrait_label}.png'


def build_background_src(sheet_name: str, scene_name: str) -> str:
    return build_story_asset_src(sheet_name, scene_name)


def build_portrait_src(portrait_label: str) -> str:
    return build_portrait_asset_src(portrait_label)


def is_cg_scene(scene_name: str) -> bool:
    return 'CG' in normalize_text(scene_name).upper()


def extract_character_name(speaker_label: str) -> str:
    text = normalize_text(speaker_label)
    if not text:
        return ''
    return text.split('-', 1)[0].strip() or text


def get_supplement_type_value(value: Any) -> int:
    text = normalize_text(value)
    if not text:
        return 0

    try:
        return int(text)
    except (TypeError, ValueError):
        return 0


def derive_effect_class(extra_text: str) -> str:
    extra = normalize_text(extra_text)
    if not extra:
        return ''

    for part in split_extra_parts(extra) or [extra]:
        normalized_part = part.replace(' ', '')
        if not any(keyword in normalized_part for keyword in SHAKE_KEYWORDS):
            continue

        if STRONG_CHAR in normalized_part:
            return 'shake-hard'
        if SOFT_CHAR in normalized_part:
            return 'shake-soft'
        return 'shake-medium'

    return ''


def derive_camera_effect(extra_text: str) -> str:
    extra = normalize_text(extra_text)
    if not extra:
        return ''

    normalized_extra = extra.replace(' ', '')
    if any(keyword in normalized_extra for keyword in PAN_RIGHT_TO_LEFT_KEYWORDS):
        return 'camera-pan-right-to-left'

    if any(keyword in normalized_extra for keyword in PAN_LEFT_TO_RIGHT_KEYWORDS):
        return 'camera-pan-left-to-right'

    for part in split_extra_parts(extra) or [extra]:
        normalized_part = part.replace(' ', '')
        if not any(keyword in normalized_part for keyword in ZOOM_KEYWORDS):
            continue

        if STRONG_CHAR in normalized_part:
            return 'camera-zoom-in-strong'
        if SOFT_CHAR in normalized_part:
            return 'camera-zoom-in-soft'
        return 'camera-zoom-in'

    return ''


def extract_tagged_extra_value(part: str, tag: str) -> str:
    normalized_part = normalize_text(part)
    if not normalized_part:
        return ''

    for separator in ('=', ':', FULL_WIDTH_COLON):
        prefix = f'{tag}{separator}'
        if normalized_part.startswith(prefix):
            return normalize_text(normalized_part[len(prefix):])

    return ''


def get_staged_actor_labels(extra_text: str) -> list[str]:
    parts = split_extra_parts(extra_text)
    if not parts:
        return []

    left_label = ''
    right_label = ''

    for part in parts:
        if not left_label:
            left_label = extract_tagged_extra_value(part, LEFT_TAG)
            if left_label:
                continue
        if not right_label:
            right_label = extract_tagged_extra_value(part, RIGHT_TAG)

    return [label for label in (left_label, right_label) if label]


def get_dual_dialogue_actor_labels(extra_text: str) -> list[str]:
    tagged = get_staged_actor_labels(extra_text)
    if tagged:
        return tagged

    return [normalize_text(part) for part in split_extra_parts(extra_text)[:2] if normalize_text(part)]


def derive_background_mode(extra_text: str, scene_name: str = '') -> str:
    extra = normalize_text(extra_text)
    normalized_scene_name = normalize_text(scene_name).replace(' ', '').strip().lower()
    if normalized_scene_name in BACKGROUND_MAIN_VALUES:
        return 'zoo-home'

    if not extra:
        return 'story'

    for part in split_extra_parts(extra) or [extra]:
        normalized_part = part.replace(' ', '').strip()
        if normalized_part.lower() in BACKGROUND_MAIN_VALUES:
            return 'zoo-home'

        if not normalized_part.startswith(BACKGROUND_KEYWORD):
            continue

        background_value = normalized_part[len(BACKGROUND_KEYWORD):]
        if background_value[:1] in ('=', ':', FULL_WIDTH_COLON):
            background_value = background_value[1:]

        if background_value.strip().lower() in BACKGROUND_MAIN_VALUES:
            return 'zoo-home'

    return 'story'


def resolve_item_reward_image_src(sheet_name: str, raw_image_value: str) -> str:
    normalized = normalize_text(raw_image_value).replace('\\', '/')
    if not normalized:
        return ''

    if re.match(r'^(?:https?:)?//', normalized):
        return normalized

    if normalized.startswith(('./', '../', '/')):
        return normalized

    has_extension = bool(re.search(r'\.[A-Za-z0-9]+$', normalized))
    if '/' in normalized:
        return normalized

    if has_extension:
        return f'./Texture/story/{sheet_name}/{normalized}'
    return build_story_asset_src(sheet_name, normalized)


def get_item_reward_title(raw_image_value: str) -> str:
    normalized = normalize_text(raw_image_value).replace('\\', '/')
    if not normalized:
        return ''

    file_name = normalized.rsplit('/', 1)[-1]
    return re.sub(r'\.[A-Za-z0-9]+$', '', file_name).strip()


def get_item_reward(sheet_name: str, extra_text: str, warnings: set[str]) -> dict[str, str] | None:
    normalized = normalize_text(extra_text)
    if not normalized:
        return None

    parts = [part.strip() for part in normalized.replace(FULL_WIDTH_COMMA, ',').split(',') if part.strip()]
    if not parts:
        return None

    raw_image_value = parts[0]
    description = FULL_WIDTH_COMMA.join(parts[1:]).strip() if len(parts) > 1 else ''
    image_src = resolve_item_reward_image_src(sheet_name, raw_image_value)
    title = get_item_reward_title(raw_image_value)

    if image_src.startswith('./Texture/story/'):
        image_path = ROOT_DIR / image_src[2:].replace('/', '\\')
        if not image_path.exists():
            warnings.add(f'Missing item reward asset: {image_path}')

    return {
        'imageSrc': image_src,
        'title': title,
        'text': description,
        'buttonLabel': '\u83b7\u5f97',
    }


def slugify_species_name(name: str) -> str:
    normalized = re.sub(r'[^a-z0-9]+', '-', name.lower())
    return normalized.strip('-') or 'species'


def get_collection_unlock(extra_text: str) -> dict[str, str] | None:
    normalized = normalize_text(extra_text)
    if not normalized:
        return None

    mapped = COLLECTION_SPECIES_MAP.get(normalized)
    if mapped:
        return { 'speciesId': mapped['speciesId'], 'speciesName': mapped['speciesName'] }

    return {
        'speciesId': slugify_species_name(normalized),
        'speciesName': normalized
    }


def resolve_cleaning_background_src(sheet_name: str, raw_value: str) -> str:
    normalized = normalize_text(raw_value).replace('\\', '/')
    if not normalized:
        return ''

    if re.match(r'^(?:https?:)?//', normalized):
        return normalized

    if normalized.startswith(('./', '../', '/')):
        return normalized

    if normalized.startswith('Texture/'):
        return f'./{normalized}'

    if re.fullmatch(r'UI_Zoo_MainBG(?:\.[A-Za-z0-9]+)?', normalized):
        zoo_asset_path = resolve_existing_asset_path(ROOT_DIR / 'Texture' / 'ZOO', normalized)
        if zoo_asset_path:
            return f'./Texture/ZOO/{zoo_asset_path.name}'
        file_name = normalized if re.search(r'\.[A-Za-z0-9]+$', normalized) else f'{normalized}.png'
        return f'./Texture/ZOO/{file_name}'

    has_extension = bool(re.search(r'\.[A-Za-z0-9]+$', normalized))
    if '/' in normalized:
        return normalized

    if has_extension:
        return f'./Texture/story/{sheet_name}/{normalized}'
    return build_story_asset_src(sheet_name, normalized)


def get_derived_half_dirty_background(dirty_background: str) -> str:
    normalized = normalize_text(dirty_background)
    if not normalized:
        return ''

    if '一半脏' in normalized:
        return ''

    if '主界面-脏' in normalized:
        return normalized.replace('主界面-脏', '主界面-一半脏')
    if '动物园-脏' in normalized:
        return normalized.replace('动物园-脏', '动物园-半脏')
    if '-脏' in normalized:
        return normalized.replace('-脏', '-一半脏')
    if '脏' in normalized:
        return normalized.replace('脏', '一半脏')
    return ''


def add_interaction_asset_warning(src: str, warnings: set[str]) -> None:
    normalized = normalize_text(src)
    if not normalized or not normalized.startswith('./'):
        return

    asset_path = ROOT_DIR / normalized[2:].replace('/', '\\')
    if not asset_path.exists():
        warnings.add(f'Missing interaction asset: {asset_path}')


def get_cleaning_interaction(
    sheet_name: str,
    background_src: str,
    extra_text: str,
    warnings: set[str],
) -> dict[str, Any]:
    normalized = normalize_text(extra_text)
    parts = [part.strip() for part in normalized.replace(FULL_WIDTH_COMMA, ',').split(',') if part.strip()] if normalized else []

    if len(parts) >= 3:
        dirty_background = resolve_cleaning_background_src(sheet_name, parts[0])
        mid_background = resolve_cleaning_background_src(sheet_name, parts[1])
        clean_background = resolve_cleaning_background_src(sheet_name, parts[2])
        prompt1 = parts[3] if len(parts) >= 4 else ''
        prompt2 = parts[4] if len(parts) >= 5 else ''
    else:
        dirty_background = normalize_text(background_src)
        mid_background = get_derived_half_dirty_background(dirty_background)
        clean_background = '' if mid_background else resolve_cleaning_background_src(sheet_name, 'UI_Zoo_MainBG')
        prompt1 = ''
        prompt2 = ''

    add_interaction_asset_warning(dirty_background, warnings)
    add_interaction_asset_warning(mid_background, warnings)
    add_interaction_asset_warning(clean_background, warnings)

    return {
        'type': 'cleaning',
        'dirtyBackground': dirty_background,
        'midBackground': mid_background,
        'cleanBackground': clean_background,
        'prompts': [prompt1, prompt2],
        'completionText': '打扫完成，动物园焕然一新！',
        'buttonLabel': '继续剧情',
    }


def get_configured_effect_class(supplement_type: int, extra_text: str) -> str:
    if supplement_type == SUPPLEMENT_TYPE_SHAKE:
        return derive_effect_class(extra_text) or 'shake-medium'
    return ''


def get_configured_camera_effect(supplement_type: int, extra_text: str) -> str:
    if supplement_type == SUPPLEMENT_TYPE_PAN_RIGHT_TO_LEFT:
        return 'camera-pan-right-to-left'
    if supplement_type == SUPPLEMENT_TYPE_ZOOM_IN:
        return derive_camera_effect(extra_text) or 'camera-zoom-in'
    return ''


def get_configured_background_mode(supplement_type: int, extra_text: str, scene_name: str = '') -> str:
    if supplement_type == SUPPLEMENT_TYPE_ZOO_HOME:
        return 'zoo-home'
    return derive_background_mode(extra_text, scene_name)


def derive_title(index_value: str, speaker_label: str, text: str, scene_name: str) -> str:
    if scene_name and is_cg_scene(scene_name):
        return scene_name

    trimmed_text = normalize_text(text)
    if trimmed_text:
        return trimmed_text[:12] + ('...' if len(trimmed_text) > 12 else '')

    speaker = normalize_text(speaker_label)
    if speaker and speaker != NARRATION_SPEAKER:
        return extract_character_name(speaker)

    if scene_name:
        return scene_name

    order = normalize_text(index_value)
    return f'第 {order or "?"} 幕'


def ensure_character(
    characters: OrderedDict[str, dict[str, Any]],
    speaker_label: str,
    warnings: set[str],
) -> tuple[str, str]:
    character_name = extract_character_name(speaker_label)
    portrait_label = normalize_text(speaker_label)

    if character_name not in characters:
        characters[character_name] = {
            'id': stable_id('character', character_name),
            'name': character_name,
            'portraitsByLabel': OrderedDict(),
        }

    portraits = characters[character_name]['portraitsByLabel']
    if portrait_label not in portraits:
        portrait_src = build_portrait_src(portrait_label)
        portrait_path = ROOT_DIR / portrait_src[2:].replace('/', '\\')
        if not portrait_path.exists():
            warnings.add(f'Missing portrait asset: {portrait_path}')

        portraits[portrait_label] = {
            'id': stable_id('portrait', portrait_label),
            'label': portrait_label,
            'src': portrait_src,
        }

    return characters[character_name]['id'], portraits[portrait_label]['id']


def convert_sheet(sheet) -> tuple[dict[str, Any] | None, list[str]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return None, []

    headers = [normalize_text(value) for value in rows[0]]
    if not any(headers):
        return None, []

    header_index = {header: index for index, header in enumerate(headers) if header}

    def cell(row_values: tuple[Any, ...], header_name: str) -> str:
        index = header_index.get(header_name)
        if index is None or index >= len(row_values):
            return ''
        return normalize_text(row_values[index])

    characters: OrderedDict[str, dict[str, Any]] = OrderedDict()
    beats: list[dict[str, Any]] = []
    warnings: set[str] = set()
    current_scene = ''

    for row_number, row_values in enumerate(rows[1:], start=2):
        order_value = cell(row_values, HEADER_ORDER)
        speaker_label = cell(row_values, HEADER_SPEAKER)
        text = cell(row_values, HEADER_TEXT)
        scene_name = cell(row_values, HEADER_SCENE)
        extra = cell(row_values, HEADER_EXTRA)
        supplement_type = get_supplement_type_value(cell(row_values, HEADER_EXTRA_TYPE))

        if scene_name:
            current_scene = scene_name
        elif current_scene:
            scene_name = current_scene

        if not any([speaker_label, text, scene_name, extra, supplement_type]):
            continue

        presentation = 'illustration' if scene_name and is_cg_scene(scene_name) else 'standard'
        effect_class = (
            get_configured_effect_class(supplement_type, extra)
            if supplement_type > 0
            else derive_effect_class(extra)
        )
        camera_effect = (
            get_configured_camera_effect(supplement_type, extra)
            if supplement_type > 0
            else derive_camera_effect(extra)
        )
        background_mode = get_configured_background_mode(supplement_type, extra, scene_name)
        background_src = '' if background_mode == 'zoo-home' else (build_background_src(sheet.title, scene_name) if scene_name else '')
        if background_src:
            background_path = ROOT_DIR / background_src[2:].replace('/', '\\')
            if not background_path.exists():
                warnings.add(f'Missing background asset: {background_path}')
        item_reward = get_item_reward(sheet.title, extra, warnings) if supplement_type == SUPPLEMENT_TYPE_ITEM_REWARD else None
        interaction = get_cleaning_interaction(sheet.title, background_src, extra, warnings) if supplement_type == SUPPLEMENT_TYPE_INTERACTION else None
        collection_unlock = get_collection_unlock(extra) if supplement_type == SUPPLEMENT_TYPE_COLLECTION else None

        if not speaker_label and not text:
            if not scene_name and not effect_class and not camera_effect and background_mode == 'story' and item_reward is None and interaction is None:
                continue

            beats.append({
                'id': stable_id('beat', f'{sheet.title}-{row_number}'),
                'title': derive_title(order_value, speaker_label, text, scene_name),
                'background': background_src,
                'type': 'narration',
                'speakerName': '',
                'text': '',
                'actorCount': 0,
                'actors': [],
                'presentation': 'transition' if scene_name and presentation != 'illustration' else presentation,
                'effectClass': effect_class,
                'cameraEffect': camera_effect,
                'backgroundMode': background_mode,
                'itemReward': item_reward,
                'interaction': interaction,
                'choices': get_option_choices(supplement_type, extra),
                'collectionUnlock': collection_unlock
            })
            continue

        is_dialogue = bool(speaker_label and speaker_label != NARRATION_SPEAKER)
        beat: dict[str, Any] = {
            'id': stable_id('beat', f'{sheet.title}-{row_number}'),
            'title': derive_title(order_value, speaker_label, text, scene_name),
            'background': background_src,
            'type': 'dialogue' if is_dialogue else 'narration',
            'speakerName': extract_character_name(speaker_label) if is_dialogue else '',
            'text': text,
            'actorCount': 0,
            'actors': [],
            'presentation': presentation,
            'effectClass': effect_class,
            'cameraEffect': camera_effect,
            'backgroundMode': background_mode,
            'itemReward': item_reward,
            'interaction': interaction,
            'choices': get_option_choices(supplement_type, extra),
            'collectionUnlock': collection_unlock
        }

        if is_dialogue:
            if supplement_type == SUPPLEMENT_TYPE_DUAL_DIALOGUE:
                actor_labels = get_dual_dialogue_actor_labels(extra)
            elif supplement_type > 0:
                actor_labels = []
            else:
                actor_labels = get_staged_actor_labels(extra)

            actor_labels = actor_labels or [speaker_label]
            actors = []
            for actor_label in actor_labels[:2]:
                normalized_actor_label = normalize_text(actor_label)
                if not normalized_actor_label:
                    continue

                character_id, portrait_id = ensure_character(characters, normalized_actor_label, warnings)
                actors.append({
                    'characterId': character_id,
                    'portraitId': portrait_id,
                })

            if not actors:
                character_id, portrait_id = ensure_character(characters, speaker_label, warnings)
                actors.append({
                    'characterId': character_id,
                    'portraitId': portrait_id,
                })

            beat['actorCount'] = len(actors)
            beat['actors'] = actors

        beats.append(beat)

    if not beats:
        return None, sorted(warnings)

    project = {
        'version': 1,
        'storyId': build_story_id(sheet.title),
        'title': sheet.title,
        'stage': {
            'maxActorsPerBeat': 2,
            'singleActorPosition': 'center',
            'doubleActorPositions': ['left', 'right'],
        },
        'characters': [
            {
                'id': character['id'],
                'name': character['name'],
                'portraits': list(character['portraitsByLabel'].values()),
            }
            for character in characters.values()
        ],
        'beats': beats,
    }
    return project, sorted(warnings)


def import_workbook(excel_path: Path) -> ImportResult:
    workbook = load_workbook(excel_path, data_only=True)
    stories: dict[str, Any] = OrderedDict()
    warnings: list[str] = []
    beat_count = 0

    for sheet in workbook.worksheets:
        project, sheet_warnings = convert_sheet(sheet)
        warnings.extend(sheet_warnings)
        if not project:
            continue
        stories[project['storyId']] = project
        beat_count += len(project.get('beats', []))

    return ImportResult(
        stories=stories,
        warnings=sorted(set(warnings)),
        story_count=len(stories),
        beat_count=beat_count,
    )


def write_generated_script(output_path: Path, result: ImportResult, excel_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(result.stories, ensure_ascii=False, indent=4)
    meta = json.dumps({
        'source': str(excel_path.name),
        'generatedAt': datetime.now(timezone.utc).astimezone().isoformat(timespec='seconds'),
        'storyCount': result.story_count,
        'beatCount': result.beat_count,
        'warnings': result.warnings,
    }, ensure_ascii=False, indent=4)

    script_text = (
        "(function initGeneratedStoryData(globalScope) {\n"
        "    'use strict';\n\n"
        f"    globalScope.WynneImportedStories = {payload};\n"
        f"    globalScope.WynneImportedStoryMeta = {meta};\n"
        "}(window));\n"
    )
    output_path.write_text(script_text, encoding='utf-8')


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Import story workbook into generated JS data.')
    parser.add_argument('--excel', type=Path, default=DEFAULT_EXCEL_PATH, help='Excel workbook path')
    parser.add_argument('--output', type=Path, default=DEFAULT_OUTPUT_PATH, help='Generated JS output path')
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    excel_path = args.excel.resolve()
    output_path = args.output.resolve()

    if not excel_path.exists():
        raise FileNotFoundError(f'Workbook not found: {excel_path}')

    result = import_workbook(excel_path)
    write_generated_script(output_path, result, excel_path)

    print(f'Imported {result.story_count} stories, {result.beat_count} beats.')
    print(f'Generated: {output_path}')
    if result.warnings:
        print('Missing assets:')
        for warning in result.warnings:
            print(f'- {warning}')

    return 0


if __name__ == '__main__':
    raise SystemExit(main())
